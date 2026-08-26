#!/usr/bin/env python3
"""Inspect the non-standard language and exchange timetable workbooks in a ZIP.

This is intentionally separate from parse_timetables.py. The language workbook
uses one cell per day and the exchange workbook has no grade/major source name,
so neither should be added to the public timetable data without a separate
mapping decision.
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SESSION_START_COLUMNS = (4, 7, 10, 13, 16, 19)
SPECIAL_MARKERS = ("english and german", "exchange")


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def parse_language_entry(line: str) -> dict[str, Any]:
    line = clean(line) or ""
    if line == "/":
        return {"kind": "empty", "raw": line}
    if line.startswith("(") and line.endswith(")"):
        return {"kind": "metadata", "raw": line}

    match = re.fullmatch(
        r"(?P<course>.+)_(?P<code>[^_]+)_(?P<teacher>[^,]+),\s*(?P<room>.+)",
        line,
    )
    if match:
        return {"kind": "class", **match.groupdict()}

    room_match = re.fullmatch(r"(?P<label>.+),\s*(?P<room>\S+)", line)
    if room_match:
        return {"kind": "unclassified", **room_match.groupdict(), "raw": line}
    return {"kind": "unclassified", "raw": line}


def split_cell_entries(value: Any) -> list[dict[str, Any]]:
    text = str(value or "")
    return [parse_language_entry(line) for line in text.splitlines() if clean(line)]


def parse_language_sheet(ws: Any) -> dict[str, Any]:
    headers = {
        column: clean(ws.cell(1, column).value) or f"column-{column}"
        for column in range(3, ws.max_column + 1)
    }
    slots = []
    for row in range(2, ws.max_row + 1):
        label = clean(ws.cell(row, 1).value)
        time = clean(ws.cell(row, 2).value)
        if not label and not time:
            continue
        days = {}
        for column, weekday in headers.items():
            value = ws.cell(row, column).value
            if value is None or clean(value) == "/":
                continue
            days[weekday] = {
                "raw": str(value),
                "entries": split_cell_entries(value),
            }
        slots.append({"label": label, "time": time, "days": days})

    return {
        "kind": "language",
        "sheet": ws.title,
        "dimensions": {"rows": ws.max_row, "columns": ws.max_column},
        "headers": list(headers.values()),
        "slots": slots,
    }


def parse_exchange_sheet(ws: Any) -> dict[str, Any]:
    sessions = []
    for index, column in enumerate(SESSION_START_COLUMNS, start=1):
        label = clean(ws.cell(1, column).value)
        time = clean(ws.cell(2, column).value)
        if label or time:
            sessions.append({"id": index, "label": label, "time": time})

    events = []
    current_week = None
    for row in range(3, ws.max_row + 1):
        if ws.cell(row, 1).value is not None:
            current_week = json_value(ws.cell(row, 1).value)
        date_label = clean(ws.cell(row, 2).value)
        weekday = clean(ws.cell(row, 3).value)
        if not date_label:
            continue
        for slot, column in enumerate(SESSION_START_COLUMNS, start=1):
            title = clean(ws.cell(row, column).value)
            teacher = clean(ws.cell(row, column + 1).value)
            room = clean(ws.cell(row, column + 2).value)
            if title or teacher or room:
                events.append(
                    {
                        "week": current_week,
                        "date": date_label,
                        "weekday": weekday,
                        "slot": slot,
                        "title": title,
                        "teacher": teacher,
                        "room": room,
                    }
                )

    catalog = []
    for row in range(1, ws.max_row + 1):
        name = clean(ws.cell(row, 23).value)
        if not name or name == "Required TH":
            continue
        catalog.append(
            {
                "name": name,
                "teacher": clean(ws.cell(row, 24).value),
                "requiredHours": json_value(ws.cell(row, 25).value),
                "requiredSessions": json_value(ws.cell(row, 26).value),
                "scheduledSessions": json_value(ws.cell(row, 27).value),
                "sourceRow": row,
            }
        )

    notices = []
    for merged in ws.merged_cells.ranges:
        if merged.min_col == 4 and merged.max_col == 21:
            notices.append(
                {
                    "range": str(merged),
                    "label": clean(ws.cell(merged.min_row, merged.min_col).value),
                }
            )

    return {
        "kind": "exchange",
        "sheet": ws.title,
        "dimensions": {"rows": ws.max_row, "columns": ws.max_column},
        "sessions": sessions,
        "events": events,
        "catalog": catalog,
        "notices": notices,
        "summary": {
            "eventCount": len(events),
            "catalogCount": len(catalog),
            "noticeCount": len(notices),
        },
    }


def inspect_workbook(filename: str, payload: bytes) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    sheets = []
    if "exchange" in filename.casefold():
        timetable_sheets = [sheet for sheet in workbook.worksheets if sheet.max_row > 2]
        sheets = [parse_exchange_sheet(sheet) for sheet in timetable_sheets]
    else:
        sheets = [parse_language_sheet(sheet) for sheet in workbook.worksheets]
    return {"file": filename, "sheets": sheets}


def select_special_files(names: list[str]) -> list[str]:
    return [
        name
        for name in names
        if Path(name).suffix.casefold() == ".xlsx"
        and any(marker in Path(name).name.casefold() for marker in SPECIAL_MARKERS)
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archive", type=Path, help="ZIP containing special timetable workbooks")
    parser.add_argument("--output", type=Path, help="write the JSON report to this path")
    args = parser.parse_args()

    with zipfile.ZipFile(args.archive) as archive:
        files = select_special_files(archive.namelist())
        if not files:
            print("error: no language or exchange workbook found", file=sys.stderr)
            return 1
        report = {
            "schemaVersion": 1,
            "sourceArchive": str(args.archive),
            "workbooks": [inspect_workbook(name, archive.read(name)) for name in files],
        }

    payload = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
        print(f"Wrote report to {args.output}")
    else:
        print(payload, end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
