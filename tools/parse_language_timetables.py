#!/usr/bin/env python3
"""Parse the term-level English and German timetable into language classes."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from data_contract import format_file, load_contract, major_name


WEEKDAYS = {
    "monday": 1,
    "tuesday": 2,
    "wednesday": 3,
    "thursday": 4,
    "friday": 5,
    "saturday": 6,
    "sunday": 7,
}
SESSION_PATTERN = re.compile(r"^(?P<slot>\d+)(?:st|nd|rd|th)\s+Session$", re.IGNORECASE)
GERMAN_PATTERN = re.compile(
    r"^German\s+(?P<section>[IVXLCDM]+)_(?P<level>[^_]+)_(?P<class_number>\d+)_\s*(?P<teachers>.+)$",
    re.IGNORECASE,
)
ENGLISH_PATTERN = re.compile(
    r"^English\s+(?P<section>[IVXLCDM]+)_(?P<class_number>\d+)_\s*(?P<teachers>.+)$",
    re.IGNORECASE,
)
ENGLISH_CATCHUP_PATTERN = re.compile(
    r"^English_Catchup_(?P<class_number>\d+)_\s*(?P<teachers>.+)$",
    re.IGNORECASE,
)
ROOM_PATTERN = re.compile(r",\s*(?P<room>[A-Za-z]\d{3})\s*$")
WEEK_RANGE_PATTERN = re.compile(r"week\s*(?P<start>\d+)\s*-\s*(?P<end>\d+)", re.IGNORECASE)
SOURCE_PATTERN = re.compile(
    r"Timetable_(?P<term>[^_]+)_\s*English and German\.xlsx",
    re.IGNORECASE,
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").replace("\u3000", " ")).strip()


def roman_to_number(value: str) -> int:
    values = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    token = value.upper()
    total = 0
    for index, character in enumerate(token):
        current = values[character]
        following = values[token[index + 1]] if index + 1 < len(token) else 0
        total += -current if current < following else current
    return total


def normalize_class_number(raw: str) -> str:
    return str(int(raw))


def normalize_level(raw: str) -> str:
    value = clean_text(raw)
    match = re.fullmatch(r"(?P<base>[A-C]\d)(?:\.\d+)?", value, re.IGNORECASE)
    return match.group("base").upper() if match else value


def parse_week_range(sheet_name: str) -> tuple[int, int] | None:
    if sheet_name.strip().casefold() == "english":
        # The workbook does not label the English range. The administrative
        # timetable remains the authoritative week mask when the web app patches it.
        return 1, 17
    match = WEEK_RANGE_PATTERN.search(sheet_name)
    if not match:
        return None
    return int(match.group("start")), int(match.group("end"))


def parse_course_line(line: str) -> dict[str, Any] | None:
    room_match = ROOM_PATTERN.search(line)
    if not room_match:
        return None
    room = room_match.group("room").upper()
    identity_and_teachers = line[: room_match.start()].strip()

    catchup_match = ENGLISH_CATCHUP_PATTERN.match(identity_and_teachers)
    if catchup_match:
        raw_number = catchup_match.group("class_number")
        teachers = [
            clean_text(item.replace("_", " "))
            for item in catchup_match.group("teachers").split(",")
            if clean_text(item.replace("_", " "))
        ]
        class_number = normalize_class_number(raw_number)
        return {
            "id": f"english:catchup:{class_number}",
            "code": f"English Catchup_{class_number}",
            "language": "English",
            "kind": "catchup",
            "section": "Catchup",
            "sectionNumber": None,
            "level": None,
            "classNumber": class_number,
            "rawClassNumber": raw_number,
            "teachers": teachers,
            "room": room,
        }

    german_match = GERMAN_PATTERN.match(identity_and_teachers)
    if german_match:
        section = german_match.group("section").upper()
        level = normalize_level(german_match.group("level"))
        raw_number = german_match.group("class_number")
        class_number = normalize_class_number(raw_number)
        teacher = clean_text(german_match.group("teachers").replace("_", " "))
        return {
            "id": f"german:{section}:{level.casefold()}:{class_number}",
            "code": f"German {section}_{level}_{class_number}",
            "language": "German",
            "kind": "standard",
            "section": section,
            "sectionNumber": roman_to_number(section),
            "level": level,
            "classNumber": class_number,
            "rawClassNumber": raw_number,
            "teachers": [teacher] if teacher else [],
            "room": room,
        }

    english_match = ENGLISH_PATTERN.match(identity_and_teachers)
    if english_match:
        section = english_match.group("section").upper()
        raw_number = english_match.group("class_number")
        class_number = normalize_class_number(raw_number)
        teacher = clean_text(english_match.group("teachers").replace("_", " "))
        return {
            "id": f"english:{section}:{class_number}",
            "code": f"English {section}_{class_number}",
            "language": "English",
            "kind": "standard",
            "section": section,
            "sectionNumber": roman_to_number(section),
            "level": None,
            "classNumber": class_number,
            "rawClassNumber": raw_number,
            "teachers": [teacher] if teacher else [],
            "room": room,
        }

    return None


def parse_workbook(path: Path, contract: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    source_match = SOURCE_PATTERN.fullmatch(path.name)
    if not source_match:
        raise ValueError(f"cannot infer term from {path.name}")
    if source_match.group("term") != contract["term"]:
        raise ValueError(
            f"workbook term {source_match.group('term')!r} does not match "
            f"contract term {contract['term']!r}"
        )

    workbook = load_workbook(path, data_only=True, read_only=False)
    classes: dict[str, dict[str, Any]] = {}
    meetings: dict[str, list[dict[str, Any]]] = defaultdict(list)
    eligibility: dict[str, set[str]] = defaultdict(set)
    warnings: list[str] = []

    for worksheet in workbook.worksheets:
        week_range = parse_week_range(worksheet.title)
        if not week_range:
            warnings.append(f"{worksheet.title}: cannot infer week range; sheet skipped")
            continue
        start_week, end_week = week_range

        headers = {
            column: clean_text(worksheet.cell(1, column).value).casefold()
            for column in range(3, worksheet.max_column + 1)
        }
        for row in range(2, worksheet.max_row + 1):
            session_label = clean_text(worksheet.cell(row, 1).value)
            session_match = SESSION_PATTERN.match(session_label)
            if not session_match:
                continue
            slot = int(session_match.group("slot"))

            for column, weekday_label in headers.items():
                weekday = WEEKDAYS.get(weekday_label)
                if not weekday:
                    continue
                cell_value = worksheet.cell(row, column).value
                if not isinstance(cell_value, str):
                    continue

                eligible_majors: list[str] = []
                for raw_line in cell_value.splitlines():
                    line = clean_text(raw_line)
                    if not line or line == "/" or re.fullmatch(r"\(\d+\s+classes\)", line, re.IGNORECASE):
                        continue
                    if line.startswith("(") and line.endswith(")"):
                        eligible_majors = [item.strip().upper() for item in line[1:-1].split("+")]
                        continue

                    parsed = parse_course_line(line)
                    if not parsed:
                        warnings.append(f"{worksheet.title}!{worksheet.cell(row, column).coordinate}: cannot parse {line!r}")
                        continue

                    class_id = parsed["id"]
                    raw_number = parsed.pop("rawClassNumber")
                    if len(raw_number) > 2 and raw_number.startswith("0"):
                        warnings.append(
                            f"{worksheet.title}!{worksheet.cell(row, column).coordinate}: "
                            f"normalized class number {raw_number!r} to {parsed['classNumber']!r}"
                        )
                    teachers = parsed.pop("teachers")
                    room = parsed.pop("room")
                    classes.setdefault(class_id, parsed)
                    eligibility[class_id].update(eligible_majors)
                    meetings[class_id].append(
                        {
                            "startWeek": start_week,
                            "endWeek": end_week,
                            "weekday": weekday,
                            "slot": slot,
                            "teachers": teachers,
                            "room": room,
                        }
                    )

    result_classes = []
    for class_id, course in classes.items():
        unique_meetings = {
            json.dumps(meeting, ensure_ascii=False, sort_keys=True): meeting
            for meeting in meetings[class_id]
        }
        result_classes.append(
            {
                **course,
                "eligibleMajorCodes": sorted(eligibility[class_id]) or None,
                "eligibleMajors": [major_name(contract, code) for code in sorted(eligibility[class_id])] or None,
                "meetings": sorted(
                    unique_meetings.values(),
                    key=lambda item: (
                        item["startWeek"],
                        item["endWeek"],
                        item["weekday"],
                        item["slot"],
                    ),
                ),
            }
        )

    result_classes.sort(
        key=lambda item: (
            item["language"],
            item["kind"],
            item["sectionNumber"] if item["sectionNumber"] is not None else 999,
            item["level"] or "",
            int(item["classNumber"]),
        )
    )
    return {
        "schemaVersion": 1,
        "source": {
            "term": source_match.group("term"),
            "file": path.name,
        },
        "classes": result_classes,
        "validation": {
            "warnings": warnings,
        },
    }, warnings


def validate_contract(classes: list[dict[str, Any]], contract: dict[str, Any]) -> None:
    expected: dict[tuple[str, str, str, str | None, str], set[str] | None] = {}
    english = contract["languages"]["english"]
    for group in english["classes"]:
        eligible = set(group["eligibleMajorCodes"])
        for class_number in group["numbers"]:
            expected[("English", "standard", english["section"], None, class_number)] = eligible
    for class_number in english["catchupClasses"]:
        expected[("English", "catchup", "Catchup", None, class_number)] = None
    for section in contract["languages"]["german"]:
        for level in section["levels"]:
            for class_number in level["classes"]:
                expected[("German", "standard", section["section"], level["level"], class_number)] = None

    actual: dict[tuple[str, str, str, str | None, str], set[str] | None] = {}
    for course in classes:
        key = (
            course["language"],
            course["kind"],
            course["section"],
            course["level"],
            course["classNumber"],
        )
        eligible = course["eligibleMajorCodes"]
        actual[key] = set(eligible) if eligible else None

    missing = sorted(set(expected) - set(actual))
    unexpected = sorted(set(actual) - set(expected))
    if missing or unexpected:
        raise ValueError(
            f"language coordinates do not match data-contract.json; "
            f"missing={missing}, unexpected={unexpected}"
        )
    mismatched_eligibility = [
        key for key in expected
        if expected[key] is not None and expected[key] != actual[key]
    ]
    if mismatched_eligibility:
        raise ValueError(
            "English eligibility does not match data-contract.json for "
            f"{mismatched_eligibility}"
        )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--contract", type=Path, default=Path("data-contract.json"))
    args = parser.parse_args()

    contract = load_contract(args.contract)

    try:
        result, warnings = parse_workbook(args.input, contract)
        validate_contract(result["classes"], contract)
    except Exception as error:  # noqa: BLE001
        print(f"ERROR {args.input.name}: {error}", file=sys.stderr)
        return 1

    term_root = args.output / contract["term"]
    term_root.mkdir(parents=True, exist_ok=True)
    for course in result["classes"]:
        kind = (
            "german"
            if course["language"] == "German"
            else "englishCatchup"
            if course["kind"] == "catchup"
            else "english"
        )
        file_name = format_file(
            contract,
            kind,
            section=course["section"],
            level=course["level"] or "",
            classNumber=course["classNumber"],
        )
        payload = {"code": course["code"], "meetings": course["meetings"]}
        (term_root / file_name).write_text(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n",
            encoding="utf-8",
        )
    print(
        f"OK {args.input.name}: classes={len(result['classes'])} "
        f"meetings={sum(len(item['meetings']) for item in result['classes'])} "
        f"warnings={len(warnings)}"
    )
    for warning in warnings:
        print(f"  warning: {warning}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
