#!/usr/bin/env python3
"""Convert the timetable workbooks in raw/ into web-friendly JSON.

The source unit is one workbook (term + grade + major). Each worksheet is
preserved as a group because worksheets with suffixes such as CS1, CS2 and CS3
contain different schedules.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from data_contract import find_major, format_file, load_contract


PARSER_VERSION = "0.2.0"
SESSION_START_COLUMNS = (4, 7, 10, 13, 16, 19)  # D, G, J, M, P, S
TIMETABLE_FIRST_ROW = 4
WEEKDAY_NUMBERS = {
    "mon": 1,
    "monday": 1,
    "tue": 2,
    "tues": 2,
    "tuesday": 2,
    "wed": 3,
    "wednesday": 3,
    "thu": 4,
    "thur": 4,
    "thurs": 4,
    "thursday": 4,
    "fri": 5,
    "friday": 5,
    "sat": 6,
    "saturday": 6,
    "sun": 7,
    "sunday": 7,
}
MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}
ROMAN_TOKEN = re.compile(r"(?<![A-Za-z0-9+#])(?P<roman>[IVXLCDM]+)(?![A-Za-z0-9+#])")


def roman_to_number(token: str) -> str:
    values = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    total = 0
    for index, character in enumerate(token):
        value = values[character]
        next_value = values[token[index + 1]] if index + 1 < len(token) else 0
        total += -value if value < next_value else value

    canonical = ""
    remaining = total
    for number, symbol in (
        (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
        (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
        (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I"),
    ):
        count, remaining = divmod(remaining, number)
        canonical += symbol * count
    return str(total) if canonical == token else token


def normalize_course_title(value: Any) -> str | None:
    text = clean(value)
    if not isinstance(text, str):
        return None

    text = ROMAN_TOKEN.sub(
        lambda match: roman_to_number(match.group("roman")),
        text,
    )
    text = re.sub(
        r"\s*\(\s*([^()]*)\s*\)\s*",
        lambda match: (
            " "
            if match.group(1).strip().upper() == "SSS"
            else f" {match.group(1).strip()} "
            if match.group(1).strip()
            else " "
        ),
        text,
    )
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line) or None


def is_self_study_title(value: Any) -> bool:
    text = clean(value)
    return isinstance(text, str) and bool(
        re.search(r"\(\s*SSS\s*\)", text, flags=re.IGNORECASE)
    )


def normalize_notice_label(value: Any) -> str | None:
    label = clean(value)
    if not isinstance(label, str):
        return None
    return "考试周" if label.casefold() in {"exam week", "exam weeks"} else label


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = re.sub(r"\s+", " ", value).strip()
        return value or None
    return value


def parse_source_name(path: Path) -> dict[str, str]:
    match = re.fullmatch(
        r"Timetable_(?P<term>[^_]+)_(?P<grade>\d{4})(?P<major>[A-Za-z]+)\.xlsx",
        path.name,
    )
    if not match:
        raise ValueError(f"cannot infer term/grade/major from {path.name}")
    return match.groupdict()


def parse_group_name(sheet_name: str, source: dict[str, str]) -> dict[str, str]:
    match = re.fullmatch(
        rf"{re.escape(source['grade'])}{re.escape(source['major'])}(?P<group>\d+)",
        sheet_name,
    )
    if match:
        group_id = match.group("group")
    else:
        # Keep an unexpected sheet name usable while reporting it to the caller.
        group_id = sheet_name
    return {"groupId": group_id, "sheetName": sheet_name}


def parse_date_text(value: Any, year: int, previous: date | None) -> date:
    text = clean(value)
    if not isinstance(text, str):
        raise ValueError(f"invalid date cell {value!r}")
    match = re.fullmatch(r"(?P<day>\d{1,2})\s+(?P<month>[A-Za-z]{3,9})\.?", text)
    if not match:
        raise ValueError(f"unrecognised date text {text!r}")
    month_name = match.group("month")[:3].lower()
    if month_name not in MONTHS:
        raise ValueError(f"unrecognised month in {text!r}")
    current = date(year, MONTHS[month_name], int(match.group("day")))
    if previous is not None and current < previous:
        current = date(year + 1, current.month, current.day)
    return current


def parse_sessions(ws: Any) -> list[dict[str, Any]]:
    sessions = []
    for index, start_column in enumerate(SESSION_START_COLUMNS, start=1):
        title = clean(ws.cell(2, start_column).value)
        time_range = clean(ws.cell(3, start_column).value)
        if not title or not time_range:
            raise ValueError(
                f"missing session header at {ws.title}!{ws.cell(2, start_column).coordinate}"
            )
        sessions.append(
            {
                "id": index,
                "label": title,
                "time": time_range,
            }
        )
    return sessions


def date_rows(ws: Any) -> tuple[int, int]:
    rows = [
        row
        for row in range(TIMETABLE_FIRST_ROW, ws.max_row + 1)
        if clean(ws.cell(row, 2).value) is not None
    ]
    if not rows:
        raise ValueError(f"{ws.title} has no timetable dates")
    return min(rows), max(rows)


def build_row_context(ws: Any, source_year: int) -> tuple[dict[int, dict[str, Any]], list[str]]:
    first_row, last_row = date_rows(ws)
    context: dict[int, dict[str, Any]] = {}
    warnings: list[str] = []
    previous_date: date | None = None
    previous_week: int | None = None

    for row in range(first_row, last_row + 1):
        try:
            current_date = parse_date_text(ws.cell(row, 2).value, source_year, previous_date)
        except ValueError as error:
            warnings.append(f"{ws.title}!B{row}: {error}")
            continue
        previous_date = current_date

        raw_week = ws.cell(row, 1).value
        if raw_week is not None:
            try:
                previous_week = int(raw_week)
            except (TypeError, ValueError):
                warnings.append(f"{ws.title}!A{row}: invalid week {raw_week!r}")
        if previous_week is None:
            warnings.append(f"{ws.title}!A{row}: missing week number")
            continue

        weekday_text = clean(ws.cell(row, 3).value)
        weekday = WEEKDAY_NUMBERS.get(str(weekday_text).lower()) if weekday_text else None
        if weekday is None:
            warnings.append(f"{ws.title}!C{row}: invalid weekday {weekday_text!r}")
            continue

        context[row] = {
            "date": current_date.isoformat(),
            "week": previous_week,
            "weekday": weekday,
        }

    return context, warnings


def merge_ranges(ws: Any) -> list[Any]:
    return [
        merged
        for merged in ws.merged_cells.ranges
        if merged.min_row >= TIMETABLE_FIRST_ROW
        and merged.min_col <= 21
        and merged.max_col >= 4
    ]


def cell_in_range(row: int, column: int, merged: Any) -> bool:
    return (
        merged.min_row <= row <= merged.max_row
        and merged.min_col <= column <= merged.max_col
    )


def parse_course_catalog(ws: Any, values_ws: Any) -> list[dict[str, Any]]:
    catalog = []
    for row in range(1, ws.max_row + 1):
        name = normalize_course_title(ws.cell(row, 23).value)  # W
        if not name or name == "Required TH":
            continue
        catalog.append(
            {
                "name": name,
                "teacher": clean(ws.cell(row, 24).value),  # X
                "requiredHours": values_ws.cell(row, 25).value,  # Y
                "requiredSessions": values_ws.cell(row, 26).value,  # Z
                "scheduledSessions": values_ws.cell(row, 27).value,  # AA
                "sourceRow": row,
            }
        )
    return catalog


def parse_group(
    ws: Any,
    values_ws: Any,
    source: dict[str, str],
    sessions: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[str]]:
    source_year = int(source["term"][:4])
    row_context, warnings = build_row_context(ws, source_year)
    merged = merge_ranges(ws)
    covered_cells = {
        (row, column)
        for item in merged
        for row in range(item.min_row, item.max_row + 1)
        for column in range(item.min_col, item.max_col + 1)
    }
    notices = []

    for item in merged:
        if item.min_col != 4 or item.max_col != 21:
            warnings.append(f"{ws.title}!{item}: unexpected timetable merge")
            continue
        label = normalize_notice_label(ws.cell(item.min_row, item.min_col).value)
        start = row_context.get(item.min_row)
        end = row_context.get(item.max_row)
        if not label or not start or not end:
            warnings.append(f"{ws.title}!{item}: cannot resolve merged notice")
            continue
        notices.append(
            {
                "kind": "notice",
                "label": label,
                "startDate": start["date"],
                "endDate": end["date"],
                "startWeek": start["week"],
                "endWeek": end["week"],
                "scope": "all-sessions",
                "sourceRange": str(item),
            }
        )

    events = []
    for row, position in row_context.items():
        for slot, start_column in enumerate(SESSION_START_COLUMNS, start=1):
            if (row, start_column) in covered_cells:
                continue
            raw_title = ws.cell(row, start_column).value
            title = normalize_course_title(raw_title)
            teacher = "自习课" if is_self_study_title(raw_title) else clean(ws.cell(row, start_column + 1).value)
            room = clean(ws.cell(row, start_column + 2).value)
            if not any(value is not None for value in (title, teacher, room)):
                continue
            if title is None:
                warnings.append(
                    f"{ws.title}!{ws.cell(row, start_column).coordinate}: "
                    "teacher/room exists without course name"
                )
                continue
            events.append(
                {
                    "kind": "course",
                    "date": position["date"],
                    "week": position["week"],
                    "weekday": position["weekday"],
                    "slot": slot,
                    "title": title,
                    "teacher": teacher,
                    "room": room,
                    "sourceCell": ws.cell(row, start_column).coordinate,
                }
            )

    catalog = parse_course_catalog(ws, values_ws)
    event_counts = Counter(event["title"] for event in events)
    for course in catalog:
        expected = course["scheduledSessions"]
        actual = event_counts[course["name"]]
        if isinstance(expected, (int, float)) and expected != actual:
            warnings.append(
                f"{ws.title}: {course['name']!r} summary scheduledSessions={expected} "
                f"but parsed events={actual}"
            )

    group_info = parse_group_name(ws.title, source)
    notices.sort(key=lambda notice: (notice["startDate"], notice["endDate"], notice["label"]))
    events.sort(key=lambda event: (event["date"], event["slot"], event["sourceCell"]))
    group = {
        **group_info,
        "events": events,
        "notices": notices,
        "courseCatalog": catalog,
        "stats": {
            "eventCount": len(events),
            "noticeCount": len(notices),
            "courseCount": len(catalog),
            "weekCount": max((item["week"] for item in row_context.values()), default=0),
        },
        "validation": {
            "warnings": warnings,
        },
    }
    return group, warnings


def parse_workbook(path: Path, contract: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    source = parse_source_name(path)
    if source["term"] != contract["term"]:
        raise ValueError(
            f"workbook term {source['term']!r} does not match contract term {contract['term']!r}"
        )
    major = find_major(contract, source["grade"], source["major"])
    workbook = load_workbook(path, data_only=False, read_only=False)
    values_workbook = load_workbook(path, data_only=True, read_only=False)
    timetable_sheets = [sheet for sheet in workbook.worksheets if sheet.max_row > 1]
    if not timetable_sheets:
        raise ValueError(f"{path.name} has no timetable worksheets")

    sessions = parse_sessions(timetable_sheets[0])
    warnings: list[str] = []
    groups = []
    for ws in timetable_sheets:
        if ws.title not in values_workbook.sheetnames:
            raise ValueError(f"missing values worksheet for {ws.title}")
        values_ws = values_workbook[ws.title]
        try:
            sheet_sessions = parse_sessions(ws)
            if sheet_sessions != sessions:
                warnings.append(f"{ws.title}: session headers differ from first worksheet")
        except ValueError as error:
            warnings.append(str(error))
        group, group_warnings = parse_group(ws, values_ws, source, sessions)
        groups.append(group)
        warnings.extend(group_warnings)

    first_event_date = min(
        (
            value
            for group in groups
            for value in [
                *(event["date"] for event in group["events"]),
                *(notice["startDate"] for notice in group["notices"]),
            ]
        ),
        default=None,
    )
    result = {
        "schemaVersion": 1,
        "source": {
            "term": source["term"],
            "grade": source["grade"],
            "majorCode": source["major"],
            "major": major["name"],
        },
        "calendar": {
            "startDate": first_event_date,
            "weekCount": max(
                (group["stats"]["weekCount"] for group in groups),
                default=0,
            ),
            "sessions": [session["time"] for session in sessions],
        },
        "groups": [
            {
                "groupId": group["groupId"],
                "events": [
                    {
                        key: value
                        for key, value in event.items()
                        if key not in {"kind", "sourceCell"}
                    }
                    for event in group["events"]
                ],
                "notices": [
                    {
                        key: value
                        for key, value in notice.items()
                        if key not in {"kind", "scope", "sourceRange"}
                    }
                    for notice in group["notices"]
                ],
            }
            for group in groups
        ],
    }
    return result, warnings


def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )


def clean_generated_json(output: Path, term: str) -> None:
    output_root = output.resolve()
    term_root = (output / term).resolve()
    if output_root not in term_root.parents:
        raise ValueError(f"refusing to clean outside output root: {term_root}")
    if term_root.exists():
        for path in term_root.glob("*.json"):
            path.unlink()
    legacy_manifest = output_root / "manifest.json"
    if legacy_manifest.exists():
        legacy_manifest.unlink()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=Path("raw"))
    parser.add_argument("--output", type=Path, default=Path("parsed"))
    parser.add_argument("--contract", type=Path, default=Path("data-contract.json"))
    parser.add_argument("--clean", action="store_true")
    args = parser.parse_args()

    contract = load_contract(args.contract)
    if args.clean:
        clean_generated_json(args.output, contract["term"])

    files = sorted(
        path
        for path in args.input.glob("*.xlsx")
        if re.fullmatch(r"Timetable_[^_]+_\d{4}[A-Za-z]+\.xlsx", path.name)
    )
    if not files:
        print(f"error: no .xlsx files found in {args.input}", file=sys.stderr)
        return 1

    total_warnings = 0
    for path in files:
        try:
            result, warnings = parse_workbook(path, contract)
        except Exception as error:  # noqa: BLE001 - report a source-specific failure
            print(f"ERROR {path.name}: {error}", file=sys.stderr)
            return 1
        source = result["source"]
        declared_major = find_major(contract, source["grade"], source["majorCode"])
        parsed_group_ids = [group["groupId"] for group in result["groups"]]
        if parsed_group_ids != declared_major["groups"]:
            print(
                f"ERROR {path.name}: parsed groups {parsed_group_ids} do not match "
                f"contract groups {declared_major['groups']}",
                file=sys.stderr,
            )
            return 1
        for group in result["groups"]:
            file_name = format_file(
                contract,
                "administrative",
                grade=source["grade"],
                majorCode=source["majorCode"],
                groupId=group["groupId"],
            )
            write_json(
                args.output / source["term"] / file_name,
                {
                    "calendar": result["calendar"],
                    "events": group["events"],
                    "notices": group["notices"],
                },
            )
        total_warnings += len(warnings)
        print(
            f"OK {path.name}: "
            f"groups={len(result['groups'])} "
            f"events={sum(len(group['events']) for group in result['groups'])} "
            f"notices={sum(len(group['notices']) for group in result['groups'])} "
            f"warnings={len(warnings)}"
        )
        for warning in warnings:
            print(f"  warning: {warning}")

    print(f"Wrote coordinate files to {args.output / contract['term']} (warnings={total_warnings})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
